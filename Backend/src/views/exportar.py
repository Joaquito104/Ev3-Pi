"""
Vistas para exportar calificaciones a PDF y Excel
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.http import HttpResponse
import csv
from io import BytesIO
from datetime import datetime
from src.models import Calificacion
from src.permissions import TieneRol

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportarPDFView(APIView):
    """
    Exportar calificaciones a PDF
    GET: /api/exportar/pdf/?estado=VALIDADA&dias=30
    """
    permission_classes = [IsAuthenticated, TieneRol]
    roles_permitidos = ["AUDITOR", "ANALISTA", "TI"]

    def get(self, request):
        if not REPORTLAB_AVAILABLE:
            return Response(
                {"detail": "ReportLab no está instalado. Instala con: pip install reportlab"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener parámetros
        estado = request.query_params.get('estado')
        dias = int(request.query_params.get('dias', 30))

        # Filtrar calificaciones
        calificaciones = Calificacion.objects.all()
        if estado:
            calificaciones = calificaciones.filter(estado=estado)

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0b1220'),
            spaceAfter=20,
            alignment=1  # Center
        )

        # Elementos del documento
        elements = []

        # Título
        title = Paragraph("📊 Reporte de Calificaciones", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))

        # Información del reporte
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        info_text = f"Generado: {fecha_actual} | Total: {calificaciones.count()} calificaciones"
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # Tabla de datos
        data = [['RUT', 'Tipo Cert.', 'Período', 'Estado', 'Auditoría', 'Creado']]

        for cal in calificaciones[:100]:  # Limitar a 100 para no sobrecargar el PDF
            auditoria = '✓' if cal.solicitar_auditoria else '-'
            data.append([
                cal.registro.rut if cal.registro else '-',
                'N/A',  # Tipo certificado no está en modelo Calificacion
                'N/A',
                cal.estado,
                auditoria,
                cal.fecha_creacion.strftime('%d/%m/%Y')
            ])

        # Estilos de tabla
        table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

        # Generar PDF
        doc.build(elements)
        buffer.seek(0)

        # Retornar respuesta
        return HttpResponse(
            buffer,
            content_type='application/pdf',
            headers={'Content-Disposition': 'attachment; filename="calificaciones.pdf"'}
        )


class ExportarExcelView(APIView):
    """
    Exportar calificaciones a Excel
    GET: /api/exportar/excel/?estado=VALIDADA&dias=30
    """
    permission_classes = [IsAuthenticated, TieneRol]
    roles_permitidos = ["AUDITOR", "ANALISTA", "TI"]

    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {"detail": "openpyxl no está instalado. Instala con: pip install openpyxl"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener parámetros
        estado = request.query_params.get('estado')

        # Filtrar calificaciones
        calificaciones = Calificacion.objects.all()
        if estado:
            calificaciones = calificaciones.filter(estado=estado)

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calificaciones"

        # Estilos
        header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados
        headers = ['RUT', 'Tipo Certificado', 'Período', 'Monto', 'Estado', 'Auditoría', 'Creado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, cal in enumerate(calificaciones, 2):
            auditoria = 'Sí' if cal.solicitar_auditoria else 'No'
            row_data = [
                cal.registro.rut if cal.registro else '',
                'N/A',
                'N/A',
                'N/A',
                cal.estado,
                auditoria,
                cal.fecha_creacion.strftime('%d/%m/%Y')
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Generar Excel
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Retornar respuesta
        return HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="calificaciones.xlsx"'}
        )


class ExportarCSVView(APIView):
    """
    Exportar calificaciones a CSV
    GET: /api/exportar/csv/?estado=VALIDADA
    """
    permission_classes = [IsAuthenticated, TieneRol]
    roles_permitidos = ["AUDITOR", "ANALISTA", "TI"]

    def get(self, request):
        estado = request.query_params.get('estado')

        calificaciones = Calificacion.objects.all()
        if estado:
            calificaciones = calificaciones.filter(estado=estado)

        # Preparar respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="calificaciones.csv"'

        writer = csv.writer(response)
        # Encabezados
        writer.writerow(['RUT', 'Tipo Certificado', 'Periodo', 'Monto', 'Estado', 'Auditoria', 'Creado'])

        for cal in calificaciones:
            auditoria = 'SI' if getattr(cal, 'solicitar_auditoria', False) else 'NO'
            writer.writerow([
                cal.registro.rut if getattr(cal, 'registro', None) else '',
                'N/A',
                'N/A',
                'N/A',
                cal.estado,
                auditoria,
                cal.fecha_creacion.strftime('%Y-%m-%d') if getattr(cal, 'fecha_creacion', None) else ''
            ])

        return response
